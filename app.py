import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
import re
import uuid

st.set_page_config(layout="wide")
st.markdown("""
    <style>
        body, .main, .block-container {
            background-color: white !important;
            color: #222 !important;
            font-family: 'Segoe UI', sans-serif;
        }
        h1, h2, h3, h4, h5, h6, .stMarkdown, .stText, .stTextLabelWrapper, label {
            color: #222 !important;
        }
        .stFileUploader, .stTextInput, .stSelectbox, .stRadio, .stButton, .stDataFrame,
        .stTextInput input, .stSelectbox div[data-baseweb="select"],
        .stSelectbox div[data-baseweb="select"] *,
        .stRadio div[role="radiogroup"] label,
        .stRadio div[role="radiogroup"] label * {
            background-color: #f5f5f5 !important;
            color: #222 !important;
            border-radius: 10px;
            font-size: 14px !important;
        }
        .stFileUploader {
            max-width: 600px !important;
            margin: 0 auto !important;
        }
        .stButton>button {
            background-color: #4CAF50;
            color: white !important;
            font-weight: bold;
            border: none;
            border-radius: 8px;
            padding: 6px 14px;
            font-size: 14px;
        }
        .stButton>button:hover {
            background-color: #45a049;
        }
        .summary-header {
            display: flex;
            font-weight: bold;
            margin-top: 1em;
            padding-bottom: 0.5rem;
            border-bottom: 2px solid #999;
            text-align: center;
            background-color: #f0f0f0;
            border-radius: 8px;
            color: #222 !important;
        }
        .summary-header div {
            flex: 1;
            padding: 0.5rem;
        }
        .number-cell {
            text-align: right !important;
            font-variant-numeric: tabular-nums;
            padding-right: 1rem;
            font-weight: bold;
            color: #222;
        }
    </style>
""", unsafe_allow_html=True)

st.title("Excel გენერატორი")

report_file = st.file_uploader("ატვირთე ანგარიშფაქტურების ფაილი (report.xlsx)", type=["xlsx"])
statement_files = st.file_uploader("ატვირთე საბანკო ამონაწერის ფაილები (statement.xlsx)", type=["xlsx"], accept_multiple_files=True)

if report_file and statement_files:
    purchases_df = pd.read_excel(report_file, sheet_name='Grid')
    
    # Process multiple bank statement files
    bank_dfs = []
    for statement_file in statement_files:
        df = pd.read_excel(statement_file)
        df['P'] = df.iloc[:, 15].astype(str).str.strip()
        df['Amount'] = pd.to_numeric(df.iloc[:, 3], errors='coerce').fillna(0)
        bank_dfs.append(df)
    
    # Combine all bank statement DataFrames
    bank_df = pd.concat(bank_dfs, ignore_index=True) if bank_dfs else pd.DataFrame()

    # Identify rows where P is missing or invalid
    missing_p_rows = bank_df[bank_df['P'].str.match(r'^\s*$') | bank_df['P'].isna()].copy()
    if not missing_p_rows.empty:
        st.warning("⚠ ამონაწერის ჩანაწერები, სადაც საიდენტიფიკაციო კოდი ვერ მოიძებნა")
        st.dataframe(missing_p_rows)

        # Allow manual entry of P codes
        for index, row in missing_p_rows.iterrows():
            manual_p = st.text_input(f"ჩაწერე საიდენტიფიკაციო კოდი ჩანაწერისთვის {index}:", key=f"manual_p_{index}")
            if manual_p and manual_p.strip():
                bank_df.loc[index, 'P'] = manual_p.strip()

    purchases_df['დასახელება'] = purchases_df['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
    purchases_df['საიდენტიფიკაციო კოდი'] = purchases_df['გამყიდველი'].apply(lambda x: ''.join(re.findall(r'\d', str(x)))[:11])

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet(title="ანგარიშფაქტურები კომპანიით")
    ws1.append(['დასახელება', 'საიდენტიფიკაციო კოდი', 'ანგარიშფაქტურების ჯამი', 'ჩარიცხული თანხა', 'სხვაობა'])

    company_summaries = []

    for company_id, group in purchases_df.groupby('საიდენტიფიკაციო კოდი'):
        company_name = group['დასახელება'].iloc[0]
        unique_invoices = group.groupby('სერია №')['ღირებულება დღგ და აქციზის ჩათვლით'].sum().reset_index()
        company_invoice_sum = unique_invoices['ღირებულება დღგ და აქციზის ჩათვლით'].sum()

        paid_sum = bank_df[bank_df["P"] == str(company_id)]["Amount"].sum()
        difference = company_invoice_sum - paid_sum

        ws1.append([company_name, company_id, company_invoice_sum, paid_sum, difference])
        company_summaries.append((company_name, company_id, company_invoice_sum, paid_sum, difference))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if 'selected_company' not in st.session_state:
        st.subheader("📋 კომპანიების ჩამონათვალი")

        search_code = st.text_input("🔎 ჩაწერე საიდენტიფიკაციო კოდი:", "")
        sort_column = st.selectbox("📊 დალაგების ველი", ["ინვოისების ჯამი", "ჩარიცხვა", "სხვაობა"])
        sort_order = st.radio("⬆️⬇️ დალაგების ტიპი", ["ზრდადობით", "კლებადობით"], horizontal=True)

        sort_index = {"ინვოისების ჯამი": 2, "ჩარიცხვა": 3, "სხვაობა": 4}[sort_column]
        reverse = sort_order == "კლებადობით"

        filtered_summaries = company_summaries
        if search_code.strip():
            filtered_summaries = [item for item in company_summaries if item[1] == search_code.strip()]

        filtered_summaries = sorted(filtered_summaries, key=lambda x: x[sort_index], reverse=reverse)

        st.markdown("""
        <div class='summary-header'>
            <div style='flex: 2;'>დასახელება</div>
            <div style='flex: 2;'>საიდენტიფიკაციო კოდი</div>
            <div style='flex: 1.5;'>ინვოისების ჯამი</div>
            <div style='flex: 1.5;'>ჩარიცხვა</div>
            <div style='flex: 1.5;'>სხვაობა</div>
        </div>
        """, unsafe_allow_html=True)

        for name, company_id, invoice_sum, paid_sum, difference in filtered_summaries:
            col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1.5])
            with col1:
                st.markdown(name)
            with col2:
                if st.button(f"{company_id}", key=f"id_{company_id}_{str(uuid.uuid4())}"):
                    st.session_state['selected_company'] = company_id
            with col3:
                st.markdown(f"<div class='number-cell'>{invoice_sum:,.2f}</div>", unsafe_allow_html=True)
            with col4:
                st.markdown(f"<div class='number-cell'>{paid_sum:,.2f}</div>", unsafe_allow_html=True)
            with col5:
                st.markdown(f"<div class='number-cell'>{difference:,.2f}</div>", unsafe_allow_html=True)

    else:
        selected_code = st.session_state['selected_company']
        df_full = pd.read_excel(report_file, sheet_name='Grid')
        df_full['დასახელება'] = df_full['გამყიდველი'].astype(str).apply(lambda x: re.sub(r'^\(\d+\)\s*', '', x).strip())
        df_full['საიდენტიფიკაციო კოდი'] = df_full['გამყიდველი'].apply(lambda x: ''.join(re.findall(r'\d', str(x)))[:11])
        matching_df = df_full[df_full['საიდენტიფიკაციო კოდი'] == selected_code]

        if not matching_df.empty:
            company_name = matching_df['დასახელება'].iloc[0]
            st.subheader(f"🔎 ({selected_code}) {company_name} - ანგარიშფაქტურები")
            st.dataframe(matching_df, use_container_width=True)

            st.subheader("🔍 მოძებნე გუგლში მასალა ან მომსახურება")
            col1, col2 = st.columns([3, 1])
            with col1:
                search_term = st.text_input("ჩაწერე სახელი ან სიტყვა:")
            with col2:
                if st.button("ძებნა"):
                    if search_term.strip():
                        search_url = f"https://www.google.com/search?q={search_term.replace(' ', '+')}"
                        st.markdown(f"[🌐 გადადი გუგლზე]({search_url})", unsafe_allow_html=True)
                    else:
                        st.warning("გთხოვ ჩაწერე ტექსტი ძებნამდე.")

            company_output = io.BytesIO()
            company_wb = Workbook()
            ws = company_wb.active
            ws.title = company_name[:31]
            ws.append(matching_df.columns.tolist())
            for row in matching_df.itertuples(index=False):
                ws.append(row)
            company_wb.save(company_output)
            company_output.seek(0)

            st.download_button(
                label=f"⬇️ ჩამოტვირთე {company_name} ინვოისების Excel",
                data=company_output,
                file_name=f"{company_name}_ინვოისები.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("📭 ჩანაწერი ვერ მოიძებნა ამ კომპანიისთვის.")

        if st.button("⬅️ დაბრუნება სრულ სიაზე"):
            del st.session_state['selected_company']
